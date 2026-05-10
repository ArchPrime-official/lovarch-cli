#!/usr/bin/env python3
"""Deliverable generators for Squad Architettura-Progetto.

Each function generates ONE deliverable type (PDF, DXF, XLSX, HTML, JSON, ZIP).
Returns bytes ready to upload to Storage.

Bucket allowed mime types reference (user-assets):
- application/pdf, image/png, image/jpeg
- application/vnd.openxmlformats-officedocument.spreadsheetml.sheet (xlsx)
- application/zip, text/plain, text/csv
- application/acad, application/x-acad, application/dwg, application/x-dwg
- (NO application/json or text/html · use text/plain)
"""
from __future__ import annotations
from io import BytesIO
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional
import json
import zipfile


# ============================================================================
# PDF · ReportLab
# ============================================================================

def gen_pdf(title: str, body_md: str, footer: str = "") -> bytes:
    """Generate PDF from title + markdown-ish body."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from reportlab.lib.enums import TA_LEFT

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm,
                              topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=20,
                                   textColor="#A16207", spaceAfter=20)
    h2_style = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=14,
                                textColor="#A16207", spaceAfter=12, spaceBefore=8)
    body_style = ParagraphStyle("body", parent=styles["BodyText"], fontSize=10,
                                  leading=14, alignment=TA_LEFT)
    footer_style = ParagraphStyle("footer", parent=styles["Italic"], fontSize=8,
                                    textColor="#666666")

    elems = [Paragraph(title, title_style), Spacer(1, 0.4*cm)]
    for line in body_md.split("\n"):
        ls = line.strip()
        if not ls:
            elems.append(Spacer(1, 0.2*cm))
        elif ls.startswith("## "):
            elems.append(Paragraph(ls[3:], h2_style))
        elif ls.startswith("### "):
            elems.append(Paragraph(f"<b>{ls[4:]}</b>", body_style))
        else:
            elems.append(Paragraph(ls.replace("**", "<b>", 1).replace("**", "</b>", 1), body_style))
    if footer:
        elems.append(Spacer(1, 1*cm))
        elems.append(Paragraph(footer, footer_style))
    doc.build(elems)
    return buf.getvalue()


# ============================================================================
# XLSX · openpyxl
# ============================================================================

def gen_xlsx(headers: List[str], rows: List[List[Any]], sheet_name: str = "Sheet1",
              col_widths: Dict[int, int] = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    fill = PatternFill(start_color="A16207", end_color="A16207", fill_type="solid")
    fnt = Font(color="FFFFFF", bold=True, size=11)
    border = Border(bottom=Side(border_style="thin", color="A16207"))
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = fill; c.font = fnt
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    for ri, row in enumerate(rows, 2):
        for ci, v in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=v)
    if col_widths:
        for col_idx, w in col_widths.items():
            ws.column_dimensions[chr(64 + col_idx)].width = w
    else:
        for col in ws.columns:
            ml = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(ml + 2, 50)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================================
# DXF · ezdxf
# ============================================================================

def gen_dxf_pianta_progetto(out_path: Path) -> int:
    """Generate pianta-progetto.dxf · 9 ambienti UNI ISO 5457."""
    import ezdxf
    from ezdxf.enums import TextEntityAlignment

    doc = ezdxf.new("R2018", setup=True)
    msp = doc.modelspace()
    layers = [
        ("MURI-NUOVI", 7), ("MURI-DEMOLIZIONE", 1), ("PORTE", 3),
        ("FINESTRE", 5), ("QUOTE", 2), ("TESTI", 6), ("ARREDI", 8),
    ]
    for name, color in layers:
        if name not in doc.layers:
            doc.layers.add(name).color = color

    # Outline + perimeter
    msp.add_lwpolyline([(0, 0), (12000, 0), (12000, 10000), (0, 10000), (0, 0)],
                        dxfattribs={"layer": "MURI-NUOVI"})
    msp.add_lwpolyline([(300, 300), (11700, 300), (11700, 9700), (300, 9700), (300, 300)],
                        dxfattribs={"layer": "MURI-NUOVI"})

    walls = [
        ((5500, 300), (5500, 4500)), ((9500, 300), (9500, 4500)),
        ((9500, 4500), (11700, 4500)), ((5500, 4500), (9500, 4500)),
        ((5500, 6500), (5500, 9700)), ((5500, 6500), (9000, 6500)),
        ((4500, 6000), (4500, 9700)), ((4500, 6000), (5500, 6000)),
        ((7500, 6500), (7500, 9700)),
    ]
    for s, e in walls:
        msp.add_line(s, e, dxfattribs={"layer": "MURI-NUOVI"})

    doors = [(1500, 300, 1900, 300), (5500, 2500, 5500, 2900),
             (5500, 5400, 5500, 5800), (4500, 7500, 4500, 7900),
             (7500, 8000, 7500, 8400), (9500, 1800, 9500, 2200)]
    for x1, y1, x2, y2 in doors:
        msp.add_line((x1, y1), (x2, y2), dxfattribs={"layer": "PORTE", "color": 3})

    windows = [(1000, 9700, 2000, 9700), (2500, 9700, 3500, 9700),
                (6000, 9700, 7000, 9700), (7500, 9700, 8500, 9700),
                (10500, 1500, 10500, 2500), (10500, 3000, 10500, 4000),
                (1500, 300, 2500, 300), (4000, 300, 5000, 300)]
    for x1, y1, x2, y2 in windows:
        msp.add_line((x1, y1), (x2, y2), dxfattribs={"layer": "FINESTRE", "color": 5})

    labels = [
        (3000, 3000, "LIVING + CUCINA"), (3000, 2500, "47.5 m²"),
        (7500, 2300, "CAMERA PADRONALE"), (7500, 1900, "18.0 m²"),
        (10600, 2000, "BAGNO PAD."), (10600, 1700, "7.0 m²"),
        (7250, 5500, "STUDIO MARCO"), (7250, 5100, "13.5 m²"),
        (2200, 8000, "CAMERA SOFIA"), (2200, 7600, "12.0 m²"),
        (6500, 8200, "BAGNO 2"), (6500, 7900, "5.5 m²"),
    ]
    for x, y, txt in labels:
        msp.add_text(txt, dxfattribs={"layer": "TESTI", "height": 200, "color": 6})\
           .set_placement((x, y), align=TextEntityAlignment.CENTER)

    # Cartiglio
    msp.add_lwpolyline([(0, -1500), (12000, -1500), (12000, -300), (0, -300), (0, -1500)],
                        dxfattribs={"layer": "MURI-NUOVI"})
    cart = [
        (200, -500, "ATTICO BRERA · Via Fiori Chiari 17 · Milano"),
        (200, -800, "Cliente: Marco Rossini & Giulia Bianchi · F.356 M.127 Sub.12"),
        (200, -1100, "Architetto: Pablo Ruan · OAPPC Milano"),
        (200, -1400, "Tav.: A.02 · pianta progetto · scala 1:50 · 25/04/2026"),
    ]
    for x, y, txt in cart:
        msp.add_text(txt, dxfattribs={"layer": "TESTI", "height": 150, "color": 6})\
           .set_placement((x, y), align=TextEntityAlignment.LEFT)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    doc.saveas(str(out_path))
    return out_path.stat().st_size


def gen_dxf_sezione(label: str, out_path: Path) -> int:
    """Generate sezione DXF (AA or BB) · simple section view."""
    import ezdxf
    from ezdxf.enums import TextEntityAlignment

    doc = ezdxf.new("R2018", setup=True)
    msp = doc.modelspace()
    for n, c in [("MURI", 7), ("QUOTE", 2), ("TESTI", 6), ("SOLAIO", 5)]:
        if n not in doc.layers:
            doc.layers.add(n).color = c

    # Section box
    msp.add_lwpolyline([(0, 0), (10000, 0), (10000, 3000), (0, 3000), (0, 0)],
                        dxfattribs={"layer": "MURI"})
    # Floor + ceiling + interior walls
    for y in [0, 200, 2700, 3000]:
        msp.add_line((0, y), (10000, y), dxfattribs={"layer": "SOLAIO"})
    for x in [3000, 6500]:
        msp.add_line((x, 200), (x, 2700), dxfattribs={"layer": "MURI"})

    msp.add_text(f"SEZIONE {label}", dxfattribs={"layer": "TESTI", "height": 250, "color": 6})\
       .set_placement((5000, 3500), align=TextEntityAlignment.CENTER)
    msp.add_text("scala 1:50 · h. 290-310 cm",
                  dxfattribs={"layer": "TESTI", "height": 150, "color": 6})\
       .set_placement((5000, -300), align=TextEntityAlignment.CENTER)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    doc.saveas(str(out_path))
    return out_path.stat().st_size


# ============================================================================
# HTML · presentazione cliente con DS V8
# ============================================================================

def gen_html_presentation(project_data: Dict[str, Any], moodboard_url: str,
                            render_urls: List[Tuple[str, str]],
                            foto_urls: Optional[List[Tuple[str, str]]] = None,
                            pinterest_urls: Optional[List[Tuple[str, str]]] = None,
                            color_palette: Optional[List[Dict]] = None,
                            docs_by_category: Optional[Dict[str, List[Dict]]] = None,
                            architect_profile: Optional[Dict[str, Any]] = None) -> str:
    """Generate FULL visual presentation HTML with DS V8.

    Sections (in order):
    - Hero · project info + key stats (chart-style cards)
    - Cliente profile · Marco + Giulia + Sofia + Otto
    - Programma spaziale · 9 ambienti grid
    - Stato attuale · galleria 6 foto (BEFORE)
    - Pinterest references · 6 immagini stile cliente
    - Concept design · moodboard composto + paleta colori (swatch)
    - Render progetto · 5 ambienti AI · BEFORE/AFTER side-by-side
    - Budget · gráfico bar split categorías
    - Timeline · gantt 90gg cantiere · SAL milestones
    - Dossier · 28 deliverables grouped by category com thumbnails
    - Imprese pre-selezionate · 3 cards
    - Footer · architetto info + contratto info
    """
    foto_urls = foto_urls or []
    pinterest_urls = pinterest_urls or []
    profile = architect_profile or {}
    brand = profile.get("brand", {})
    fiscal = profile.get("fiscal", {})
    creds = profile.get("italian_credentials", {})
    style_data = profile.get("style", {})

    # Architect derived strings
    arch_name_h = profile.get("full_name") or "[Architetto]"
    arch_company_h = profile.get("company_name") or brand.get("name") or ""
    arch_logo = brand.get("logo_url", "")
    arch_email_h = profile.get("email", "")
    arch_phone_h = profile.get("phone", "")
    arch_pec_h = creds.get("pec", "")
    arch_ordine_h = creds.get("ordine_professionale", "")
    arch_matricola_h = creds.get("matricola", "")
    arch_address_h = ""
    if fiscal.get("street"):
        arch_address_h = fiscal["street"]
        if fiscal.get("city"):
            arch_address_h += f", {fiscal.get('postal_code', '')} {fiscal['city']}"
        if fiscal.get("country"):
            arch_address_h += f" ({fiscal['country']})"

    color_palette = color_palette or [
        {"hex": "#D4CBC0", "name": "Beige caldo", "usage": "Pareti", "percentage": 30},
        {"hex": "#B8865A", "name": "Terra di siena", "usage": "Accenti", "percentage": 20},
        {"hex": "#87A47A", "name": "Verde salvia", "usage": "Sofia", "percentage": 15},
        {"hex": "#C8A296", "name": "Rosa antico", "usage": "Tessili", "percentage": 10},
        {"hex": "#F2EDE4", "name": "Crema", "usage": "Soffitti", "percentage": 15},
        {"hex": "#5C7B6F", "name": "Verde profondo", "usage": "Studio", "percentage": 10},
    ]
    docs_by_category = docs_by_category or {}

    foto_cards = "\n".join(
        f'<div class="gallery-card"><img src="{url}" alt="{name}" loading="lazy"/><div class="caption">{name}</div></div>'
        for name, url in foto_urls
    ) if foto_urls else '<p class="muted">Foto stato attuale non disponibili</p>'

    pin_cards = "\n".join(
        f'<div class="gallery-card"><img src="{url}" alt="{name}" loading="lazy"/><div class="caption">{name}</div></div>'
        for name, url in pinterest_urls
    ) if pinterest_urls else '<p class="muted">References non disponibili</p>'

    render_cards = "\n".join(
        f'<div class="render-card"><img src="{url}" alt="{ambiente}" loading="lazy"/>'
        f'<div class="caption">{ambiente.replace("-progetto", "").replace("-", " ").title()}</div></div>'
        for ambiente, url in render_urls
    )

    # Before/after pairs (match foto with corresponding render by ambiente)
    foto_by_room = {}
    for name, url in foto_urls:
        # 03-soggiorno.jpg → soggiorno
        slug = name.split(".")[0].split("-", 1)[-1] if "-" in name else name
        foto_by_room[slug] = (name, url)

    ba_pairs = []
    for ambiente, render_url in render_urls:
        room_slug = ambiente.replace("-progetto", "")
        if room_slug in foto_by_room:
            foto_name, foto_url = foto_by_room[room_slug]
            ba_pairs.append((room_slug, foto_url, render_url))

    ba_cards = "\n".join(f"""
    <div class="ba-card">
      <div class="ba-header">{slug.replace("-", " ").title()}</div>
      <div class="ba-grid">
        <div class="ba-side"><span class="ba-tag tag-before">PRIMA · 1980s</span><img src="{f}" alt="before"/></div>
        <div class="ba-side"><span class="ba-tag tag-after">DOPO · progetto</span><img src="{r}" alt="after"/></div>
      </div>
    </div>""" for slug, f, r in ba_pairs)

    palette_swatches = "\n".join(
        f'<div class="swatch"><div class="swatch-color" style="background:{c["hex"]}"></div>'
        f'<div class="swatch-info"><div class="swatch-name">{c["name"]}</div>'
        f'<div class="swatch-meta">{c["hex"]} · {c["percentage"]}%</div>'
        f'<div class="swatch-usage">{c.get("usage", "")}</div></div></div>'
        for c in color_palette
    )

    # Budget split chart (CSS bars)
    budget_categories = [
        ("Opere edili", 30808, "#A16207"),
        ("Impianti", 40175, "#B8865A"),
        ("Finiture", 48310, "#87A47A"),
        ("Bagni e cucina", 47610, "#C8A296"),
        ("Sicurezza + IVA", 13097, "#5C7B6F"),
    ]
    total_budget = sum(v for _, v, _ in budget_categories)
    budget_bars = "\n".join(
        f'<div class="bar-row"><div class="bar-label">{n}</div>'
        f'<div class="bar-track"><div class="bar-fill" style="width:{v/total_budget*100:.1f}%;background:{col}"></div>'
        f'<div class="bar-value">€ {v:,.0f}</div></div></div>'.replace(",", ".")
        for n, v, col in budget_categories
    )

    # Timeline gantt
    phases_timeline = [
        ("Progettazione", "25/04", "05/06", 26.7, "#A16207"),
        ("Pratiche edilizie", "06/06", "30/06", 16.7, "#B8865A"),
        ("Demolizioni", "01/07", "10/07", 6.7, "#5C7B6F"),
        ("Impianti", "11/07", "31/07", 14.0, "#87A47A"),
        ("Tramezze + intonaci", "01/08", "20/08", 13.3, "#C8A296"),
        ("Pavimenti + restauri", "21/08", "10/09", 14.0, "#A16207"),
        ("Cucina + bagni", "11/09", "30/09", 13.3, "#B8865A"),
        ("Verniciature + finiture", "01/10", "20/10", 13.3, "#87A47A"),
        ("Collaudo + consegna", "21/10", "31/10", 7.3, "#5C7B6F"),
    ]
    timeline_rows = "\n".join(
        f'<div class="tl-row"><div class="tl-label">{n}</div><div class="tl-bar" style="background:{col};width:{w}%">'
        f'<span>{s} → {e}</span></div></div>'
        for n, s, e, w, col in phases_timeline
    )

    sal_milestones = """
    <div class="sal-row"><div class="sal-marker">15%</div><div><strong>SAL 1 · Firma contratto</strong> · 25/04/2026 · € 3.300</div></div>
    <div class="sal-row"><div class="sal-marker">25%</div><div><strong>SAL 2 · CILA depositata</strong> · 12/06/2026 · € 5.500</div></div>
    <div class="sal-row"><div class="sal-marker">25%</div><div><strong>SAL 3 · 50% lavori</strong> · 15/08/2026 · € 5.500</div></div>
    <div class="sal-row"><div class="sal-marker">35%</div><div><strong>SAL 4 · Consegna chiavi</strong> · 31/10/2026 · € 7.700</div></div>
    """

    # Documents grouped
    cat_meta = {
        "cliente": ("👤", "Per il Cliente", "#CA8A04"),
        "comune": ("🏛", "Per il Comune", "#A855F7"),
        "impresa": ("👷", "Per l'Impresa", "#F59E0B"),
        "studio": ("📋", "Studio interno", "#10B981"),
        "ingegneri": ("⚙", "Ingegneri", "#64748B"),
        "altro": ("📁", "Altri", "#A1A1AA"),
    }
    cat_sections = ""
    for cat, files in docs_by_category.items():
        if not files:
            continue
        icon, label, color = cat_meta.get(cat, ("📁", cat.title(), "#A1A1AA"))
        items = "\n".join(
            f'<a href="{f.get("public_url", "#")}" target="_blank" class="doc-item">'
            f'<span class="doc-name">{f["name"]}</span>'
            f'<span class="doc-meta">{f.get("size", 0) // 1024} KB</span></a>'
            for f in files
        )
        cat_sections += f"""
        <div class="cat-card" style="border-top: 4px solid {color}">
          <div class="cat-header">
            <span class="cat-icon">{icon}</span>
            <h3>{label}</h3>
            <span class="cat-count">{len(files)}</span>
          </div>
          <div class="cat-items">{items}</div>
        </div>"""

    return f"""<!DOCTYPE html>
<html lang="it"><head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Attico Brera · Presentazione · Marco Rossini & Giulia Bianchi</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;700;900&family=Outfit:wght@400;500;600;700&family=DM+Sans:wght@400;500;600;700&family=Inter:wght@400;500;600&display=swap" rel="stylesheet">
<style>
:root {{
  --bg: #FAF9F7;
  --card: #FFFFFF;
  --text: #18181B;
  --text-muted: #71717A;
  --gold: #A16207;
  --gold-light: #CA8A04;
  --gold-bg: rgba(161,98,7,0.08);
  --emerald: #10B981;
  --amber: #F59E0B;
  --red: #EF4444;
  --border: #E5E5E5;
}}
* {{ box-sizing: border-box; }}
body {{ background: var(--bg); color: var(--text); font-family: 'Inter', sans-serif; margin: 0; line-height: 1.5; }}
.font-display {{ font-family: 'Playfair Display', serif; }}
.font-outfit {{ font-family: 'Outfit', sans-serif; }}
.font-dm {{ font-family: 'DM Sans', sans-serif; }}
.muted {{ color: var(--text-muted); }}

/* HERO */
.hero {{
  background: linear-gradient(135deg, #18181B 0%, #2A1810 50%, #A16207 100%);
  color: white;
  padding: 80px 40px 60px;
  position: relative;
  overflow: hidden;
}}
.hero::before {{
  content: '';
  position: absolute;
  inset: 0;
  background-image: radial-gradient(circle at 20% 50%, rgba(255,255,255,0.05) 0%, transparent 30%);
}}
.hero-inner {{ max-width: 1200px; margin: 0 auto; position: relative; }}
.hero .chip {{
  display: inline-block; padding: 6px 16px; background: rgba(255,255,255,0.12);
  border: 1px solid rgba(255,255,255,0.2); border-radius: 999px;
  font-family: 'DM Sans'; font-size: 13px; font-weight: 500; margin-bottom: 24px;
}}
.hero h1 {{ font-family: 'Playfair Display', serif; font-size: 64px; margin: 0; line-height: 1; font-weight: 700; }}
.hero h1 em {{ color: #FCD34D; font-style: italic; font-weight: 400; }}
.hero .subtitle {{ margin-top: 16px; opacity: 0.85; font-size: 18px; }}
.hero-stats {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; margin-top: 48px; }}
.stat-card {{ background: rgba(255,255,255,0.08); border: 1px solid rgba(255,255,255,0.15); border-radius: 12px; padding: 20px; backdrop-filter: blur(10px); }}
.stat-num {{ font-family: 'DM Sans'; font-size: 36px; font-weight: 700; color: #FCD34D; }}
.stat-label {{ font-size: 13px; opacity: 0.75; margin-top: 4px; }}

/* SECTION */
section.block {{ padding: 60px 40px; max-width: 1200px; margin: 0 auto; }}
.sec-title {{ display: flex; align-items: baseline; gap: 16px; margin-bottom: 32px; padding-bottom: 16px; border-bottom: 1px solid var(--border); }}
.sec-num {{ font-family: 'DM Sans'; font-size: 13px; color: var(--gold); font-weight: 700; letter-spacing: 0.1em; }}
.sec-title h2 {{ font-family: 'Playfair Display', serif; font-size: 36px; margin: 0; line-height: 1; }}
.sec-intro {{ font-size: 16px; color: var(--text-muted); max-width: 720px; margin-bottom: 32px; }}

/* CLIENT PROFILE */
.profile-grid {{ display: grid; grid-template-columns: 1fr 2fr; gap: 24px; }}
.profile-card {{ background: var(--card); padding: 28px; border-radius: 16px; border: 1px solid var(--border); }}
.profile-avatar {{ width: 72px; height: 72px; border-radius: 50%; background: linear-gradient(135deg, var(--gold) 0%, var(--gold-light) 100%); color: white; display: flex; align-items: center; justify-content: center; font-family: 'Playfair Display'; font-size: 32px; font-weight: 700; margin-bottom: 20px; }}
.profile-row {{ display: flex; justify-content: space-between; padding: 8px 0; border-bottom: 1px solid var(--border); font-size: 14px; }}
.profile-row:last-child {{ border-bottom: none; }}
.profile-row .label {{ color: var(--text-muted); }}
.profile-row .value {{ font-family: 'DM Sans'; font-weight: 500; }}
.quote-card {{ background: linear-gradient(135deg, var(--gold-bg) 0%, transparent 100%); padding: 32px; border-left: 4px solid var(--gold); border-radius: 0 16px 16px 0; }}
.quote-card .quote {{ font-family: 'Playfair Display'; font-size: 22px; font-style: italic; line-height: 1.4; }}
.quote-card .source {{ font-size: 13px; color: var(--text-muted); margin-top: 16px; }}

/* PROGRAMMA */
.programma-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(220px, 1fr)); gap: 16px; }}
.amb-card {{ background: var(--card); border: 1px solid var(--border); border-radius: 12px; padding: 18px; transition: all 0.3s; }}
.amb-card:hover {{ transform: translateY(-2px); box-shadow: 0 8px 24px rgba(0,0,0,0.08); }}
.amb-name {{ font-family: 'DM Sans'; font-weight: 600; color: var(--gold); margin-bottom: 8px; }}
.amb-area {{ font-size: 24px; font-family: 'DM Sans'; font-weight: 700; }}
.amb-desc {{ font-size: 12px; color: var(--text-muted); margin-top: 8px; line-height: 1.4; }}

/* GALLERY */
.gallery-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(280px, 1fr)); gap: 16px; }}
.gallery-card {{ background: var(--card); border: 1px solid var(--border); border-radius: 12px; overflow: hidden; }}
.gallery-card img {{ width: 100%; aspect-ratio: 4/3; object-fit: cover; display: block; }}
.gallery-card .caption {{ padding: 12px 14px; font-size: 13px; font-family: 'DM Sans'; font-weight: 500; }}

/* MOODBOARD */
.moodboard-section {{ display: grid; grid-template-columns: 2fr 1fr; gap: 24px; align-items: start; }}
.moodboard-img {{ width: 100%; border-radius: 16px; box-shadow: 0 12px 40px rgba(0,0,0,0.12); }}
.palette-list {{ display: grid; gap: 12px; }}
.swatch {{ display: flex; align-items: center; gap: 14px; padding: 12px; background: var(--card); border: 1px solid var(--border); border-radius: 10px; }}
.swatch-color {{ width: 48px; height: 48px; border-radius: 8px; flex-shrink: 0; box-shadow: inset 0 0 0 1px rgba(0,0,0,0.05); }}
.swatch-name {{ font-family: 'DM Sans'; font-weight: 600; font-size: 13px; }}
.swatch-meta {{ font-size: 11px; color: var(--text-muted); margin-top: 2px; }}
.swatch-usage {{ font-size: 11px; color: var(--text-muted); margin-top: 2px; font-style: italic; }}

/* BEFORE/AFTER */
.ba-grid-wrap {{ display: grid; gap: 24px; }}
.ba-card {{ background: var(--card); border: 1px solid var(--border); border-radius: 16px; overflow: hidden; }}
.ba-header {{ padding: 14px 20px; background: linear-gradient(90deg, var(--gold-bg) 0%, transparent 100%); font-family: 'Outfit'; font-weight: 600; font-size: 16px; }}
.ba-grid {{ display: grid; grid-template-columns: 1fr 1fr; }}
.ba-side {{ position: relative; }}
.ba-side img {{ width: 100%; aspect-ratio: 4/3; object-fit: cover; display: block; }}
.ba-tag {{ position: absolute; top: 12px; left: 12px; padding: 4px 10px; border-radius: 6px; font-family: 'DM Sans'; font-size: 11px; font-weight: 600; letter-spacing: 0.05em; }}
.tag-before {{ background: rgba(0,0,0,0.7); color: white; }}
.tag-after {{ background: var(--gold); color: white; }}

/* RENDERS */
.renders-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(280px, 1fr)); gap: 16px; }}
.render-card {{ background: var(--card); border: 1px solid var(--border); border-radius: 12px; overflow: hidden; }}
.render-card img {{ width: 100%; aspect-ratio: 1; object-fit: cover; display: block; }}
.render-card .caption {{ padding: 12px 14px; font-family: 'DM Sans'; font-weight: 500; font-size: 14px; }}

/* BUDGET */
.budget-card {{ background: var(--card); border: 1px solid var(--border); border-radius: 16px; padding: 32px; }}
.budget-total {{ font-family: 'Playfair Display'; font-size: 48px; font-weight: 700; color: var(--gold); }}
.budget-total-label {{ color: var(--text-muted); font-size: 14px; margin-bottom: 24px; }}
.bar-row {{ display: grid; grid-template-columns: 200px 1fr; gap: 16px; padding: 10px 0; align-items: center; }}
.bar-label {{ font-family: 'DM Sans'; font-size: 14px; }}
.bar-track {{ background: var(--bg); border-radius: 6px; height: 32px; position: relative; overflow: hidden; }}
.bar-fill {{ height: 100%; border-radius: 6px; transition: width 0.6s ease; }}
.bar-value {{ position: absolute; right: 12px; top: 50%; transform: translateY(-50%); font-family: 'DM Sans'; font-weight: 600; font-size: 13px; }}

/* TIMELINE */
.timeline-card {{ background: var(--card); border: 1px solid var(--border); border-radius: 16px; padding: 32px; }}
.tl-row {{ display: grid; grid-template-columns: 200px 1fr; gap: 16px; padding: 8px 0; align-items: center; }}
.tl-label {{ font-family: 'DM Sans'; font-size: 13px; }}
.tl-bar {{ height: 24px; border-radius: 4px; padding: 0 12px; display: flex; align-items: center; color: white; font-family: 'DM Sans'; font-size: 11px; font-weight: 500; }}
.sal-grid {{ display: grid; grid-template-columns: repeat(2, 1fr); gap: 16px; margin-top: 32px; }}
.sal-row {{ background: var(--bg); padding: 16px; border-radius: 10px; display: flex; align-items: center; gap: 16px; }}
.sal-marker {{ width: 56px; height: 56px; border-radius: 50%; background: var(--gold); color: white; display: flex; align-items: center; justify-content: center; font-family: 'Playfair Display'; font-weight: 700; font-size: 18px; flex-shrink: 0; }}

/* DOCS */
.docs-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(280px, 1fr)); gap: 16px; }}
.cat-card {{ background: var(--card); border: 1px solid var(--border); border-radius: 12px; overflow: hidden; }}
.cat-header {{ padding: 16px 20px; display: flex; align-items: center; gap: 12px; border-bottom: 1px solid var(--border); }}
.cat-icon {{ font-size: 20px; }}
.cat-header h3 {{ font-family: 'Outfit'; font-weight: 600; font-size: 15px; margin: 0; flex: 1; }}
.cat-count {{ font-family: 'DM Sans'; background: var(--bg); padding: 2px 10px; border-radius: 999px; font-size: 12px; font-weight: 600; }}
.cat-items {{ padding: 8px; }}
.doc-item {{ display: flex; justify-content: space-between; padding: 8px 12px; border-radius: 6px; font-size: 13px; text-decoration: none; color: var(--text); transition: background 0.2s; }}
.doc-item:hover {{ background: var(--bg); }}
.doc-item .doc-name {{ font-family: 'DM Sans'; }}
.doc-item .doc-meta {{ color: var(--text-muted); font-family: 'DM Sans'; font-size: 11px; }}

/* IMPRESE */
.imp-grid {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 16px; }}
.imp-card {{ background: var(--card); border: 1px solid var(--border); border-radius: 12px; padding: 24px; }}
.imp-name {{ font-family: 'DM Sans'; font-weight: 700; margin-bottom: 8px; }}
.imp-meta {{ font-size: 13px; color: var(--text-muted); }}

/* FOOTER */
footer {{ background: var(--text); color: white; padding: 60px 40px; margin-top: 60px; }}
footer .inner {{ max-width: 1200px; margin: 0 auto; }}
footer h3 {{ font-family: 'Playfair Display'; font-size: 28px; margin: 0 0 16px; }}
footer .links {{ display: flex; gap: 24px; flex-wrap: wrap; margin-top: 24px; opacity: 0.75; font-size: 13px; }}

@media (max-width: 768px) {{
  .hero h1 {{ font-size: 40px; }}
  .hero-stats {{ grid-template-columns: repeat(2, 1fr); }}
  .profile-grid, .moodboard-section {{ grid-template-columns: 1fr; }}
  .imp-grid {{ grid-template-columns: 1fr; }}
  .ba-grid {{ grid-template-columns: 1fr; }}
  .sal-grid {{ grid-template-columns: 1fr; }}
}}
</style>
</head><body>

<section class="hero">
  <div class="hero-inner">
    <div class="chip">📍 Milano · Brera · 25 Aprile 2026</div>
    <h1>Attico <em>Brera</em></h1>
    <div class="subtitle">Marco Rossini & Giulia Bianchi · Via Fiori Chiari 17, 20121 Milano</div>
    <div class="hero-stats">
      <div class="stat-card"><div class="stat-num">120</div><div class="stat-label">m² · superficie lorda</div></div>
      <div class="stat-card"><div class="stat-num">€180K</div><div class="stat-label">budget lavori (IVA inclusa)</div></div>
      <div class="stat-card"><div class="stat-num">90</div><div class="stat-label">giorni cantiere</div></div>
      <div class="stat-card"><div class="stat-num">9</div><div class="stat-label">ambienti programma spaziale</div></div>
    </div>
  </div>
</section>

<section class="block" id="cliente">
  <div class="sec-title">
    <span class="sec-num">01 · CLIENTE</span>
    <h2>Famiglia Rossini-Bianchi</h2>
  </div>
  <div class="profile-grid">
    <div class="profile-card">
      <div class="profile-avatar">MR</div>
      <div style="font-family: 'Playfair Display'; font-size: 22px; font-weight: 600;">Marco & Giulia</div>
      <div class="muted" style="font-size: 13px; margin-bottom: 16px;">+ Sofia (6) · Otto Labrador</div>
      <div class="profile-row"><span class="label">Marco</span><span class="value">42 · Avvocato d'affari</span></div>
      <div class="profile-row"><span class="label">Giulia</span><span class="value">38 · Designer gioielli</span></div>
      <div class="profile-row"><span class="label">Reddito</span><span class="value">€180K/anno</span></div>
      <div class="profile-row"><span class="label">PEC</span><span class="value" style="font-size: 11px;">rossini.bianchi@pec.it</span></div>
      <div class="profile-row"><span class="label">Stato</span><span class="value" style="color: var(--emerald);">✓ Contratto firmato</span></div>
    </div>
    <div class="quote-card">
      <p class="quote">"Vogliamo uno spazio che respira. La cucina deve essere il centro — cuciniamo tutti i giorni, riceviamo amici. Sofia ha bisogno di una stanza sua che cresca con lei. Amiamo il legno chiaro, il travertino, le tinte terra. <strong>Non vogliamo il total white milanese.</strong>"</p>
      <div class="source">— Marco Rossini · riunione 22/04/2026</div>
    </div>
  </div>
</section>

<section class="block" id="programma">
  <div class="sec-title">
    <span class="sec-num">02 · PROGRAMMA SPAZIALE</span>
    <h2>9 ambienti · 120 m²</h2>
  </div>
  <div class="programma-grid">
    <div class="amb-card"><div class="amb-name">Living open-space</div><div class="amb-area">47.5 m²</div><div class="amb-desc">Soggiorno + cucina + dining · isola centrale · seminato preservato</div></div>
    <div class="amb-card"><div class="amb-name">Camera padronale</div><div class="amb-area">18 m²</div><div class="amb-desc">Letto king + cabina armadio walk-in · soffitto decorato originale</div></div>
    <div class="amb-card"><div class="amb-name">Studio Marco</div><div class="amb-area">13.5 m²</div><div class="amb-desc">Vetrata interna · libreria · porta scorrevole acustica</div></div>
    <div class="amb-card"><div class="amb-name">Camera Sofia</div><div class="amb-area">12 m²</div><div class="amb-desc">Carta da parati botanica · scrivania · materiali sani</div></div>
    <div class="amb-card"><div class="amb-name">Bagno padronale</div><div class="amb-area">7 m²</div><div class="amb-desc">Spa-like · gres travertino · doccia walk-in · cromoterapia</div></div>
    <div class="amb-card"><div class="amb-name">Bagno secondo</div><div class="amb-area">5.5 m²</div><div class="amb-desc">Vasca · lavabo · WC · per Sofia + ospiti</div></div>
    <div class="amb-card"><div class="amb-name">Ingresso</div><div class="amb-area">6 m²</div><div class="amb-desc">Funzionale · scarpe + cappotti · custom oak</div></div>
    <div class="amb-card"><div class="amb-name">Lavanderia</div><div class="amb-area">3.5 m²</div><div class="amb-desc">Lavatrice + asciugatrice + spazio Otto</div></div>
    <div class="amb-card"><div class="amb-name">Terrazzo</div><div class="amb-area">20 m²</div><div class="amb-desc">SW · BBQ · accessibile da living E camera padronale</div></div>
  </div>
</section>

<section class="block" id="stato-attuale">
  <div class="sec-title">
    <span class="sec-num">03 · STATO ATTUALE</span>
    <h2>Foto degli ambienti · prima della ristrutturazione</h2>
  </div>
  <p class="sec-intro">Edificio del 1910 con ristrutturazione anni '80 non funzionale. Da preservare: soffitti decorati originali e seminato veneziano nel soggiorno.</p>
  <div class="gallery-grid">{foto_cards}</div>
</section>

<section class="block" id="references">
  <div class="sec-title">
    <span class="sec-num">04 · STILE PREFERENZIALE</span>
    <h2>References del cliente</h2>
  </div>
  <p class="sec-intro">12 immagini Pinterest selezionate dal cliente: Vincenzo De Cotiis, Studiopepe, Marcante & Testa, Hotel Vilòn, Cassina, Devon&Devon. Hashtag: #materialhonesty #wabisabi #neoclassicocontemporaneo</p>
  <div class="gallery-grid">{pin_cards}</div>
</section>

<section class="block" id="moodboard">
  <div class="sec-title">
    <span class="sec-num">05 · CONCEPT</span>
    <h2>Moodboard + paleta colori</h2>
  </div>
  <p class="sec-intro">Composizione editoriale di materiali, finiture, paleta cromatica · base concept progettuale.</p>
  <div class="moodboard-section">
    <img src="{moodboard_url}" alt="Moodboard composto" class="moodboard-img"/>
    <div class="palette-list">{palette_swatches}</div>
  </div>
</section>

<section class="block" id="before-after" style="background: linear-gradient(135deg, var(--gold-bg) 0%, transparent 100%); border-radius: 24px; max-width: 1280px;">
  <div class="sec-title">
    <span class="sec-num">06 · BEFORE / AFTER</span>
    <h2>Lo spazio · prima e dopo</h2>
  </div>
  <p class="sec-intro">Render generati con AI image-to-image · <strong>preservando la struttura originale</strong> (muri, finestre, soffitti decorati) e aggiungendo lo stile post-ristrutturazione.</p>
  <div class="ba-grid-wrap">{ba_cards}</div>
</section>

<section class="block" id="render">
  <div class="sec-title">
    <span class="sec-num">07 · RENDER PROGETTO</span>
    <h2>5 ambienti · post-ristrutturazione</h2>
  </div>
  <div class="renders-grid">{render_cards}</div>
</section>

<section class="block" id="budget">
  <div class="sec-title">
    <span class="sec-num">08 · BUDGET</span>
    <h2>€180K · breakdown per categoria</h2>
  </div>
  <div class="budget-card">
    <div class="budget-total">€ 180.000</div>
    <div class="budget-total-label">Totale lavori · IVA 10% prima casa inclusa · Onorari professionali separati € 22.000 (12,2%)</div>
    {budget_bars}
  </div>
</section>

<section class="block" id="timeline">
  <div class="sec-title">
    <span class="sec-num">09 · TIMELINE</span>
    <h2>90 giorni cantiere · 1 lug → 31 ott 2026</h2>
  </div>
  <div class="timeline-card">
    <div style="margin-bottom: 24px;">
      <div style="font-family: 'Outfit'; font-weight: 600; margin-bottom: 12px;">Cronoprogramma per fase</div>
      {timeline_rows}
    </div>
    <div>
      <div style="font-family: 'Outfit'; font-weight: 600; margin-top: 32px; margin-bottom: 12px;">SAL milestones · pagamenti onorari</div>
      <div class="sal-grid">{sal_milestones}</div>
    </div>
  </div>
</section>

<section class="block" id="dossier">
  <div class="sec-title">
    <span class="sec-num">10 · DOSSIER</span>
    <h2>Tutti i deliverables · per categoria</h2>
  </div>
  <p class="sec-intro">Pacchetto completo generato dallo squad architettura-progetto · pronto per firma del professionista, deposito al Comune, gara impresa.</p>
  <div class="docs-grid">{cat_sections}</div>
</section>

<section class="block" id="imprese">
  <div class="sec-title">
    <span class="sec-num">11 · IMPRESE</span>
    <h2>3 imprese pre-selezionate dal cliente</h2>
  </div>
  <div class="imp-grid">
    <div class="imp-card">
      <div class="imp-name">Edilcasa Lombardia Srl</div>
      <div class="imp-meta">📍 Buccinasco<br>Esperienza Brera · raccomandata da amico avvocato</div>
    </div>
    <div class="imp-card">
      <div class="imp-name">Costruzioni Galimberti</div>
      <div class="imp-meta">📍 Milano · Via Marghera<br>Ha rifatto l'appartamento dei genitori di Giulia</div>
    </div>
    <div class="imp-card">
      <div class="imp-name">Restauri Brambilla & Co.</div>
      <div class="imp-meta">📍 Cernusco<br>Specialisti edifici storici · più caro · garanzia restauro</div>
    </div>
  </div>
</section>

<section class="block" id="studio" style="background: linear-gradient(180deg, var(--bg) 0%, white 100%); border-radius: 24px;">
  <div class="sec-title">
    <span class="sec-num">12 · LO STUDIO</span>
    <h2>{arch_company_h or arch_name_h}</h2>
  </div>
  {f'<img src="{arch_logo}" alt="{arch_company_h} logo" style="max-height: 80px; margin-bottom: 24px;"/>' if arch_logo else ''}
  <div class="profile-grid">
    <div class="profile-card">
      <div class="profile-avatar">{arch_name_h[:2].upper() if arch_name_h else 'AR'}</div>
      <div style="font-family: 'Playfair Display'; font-size: 22px; font-weight: 600;">Arch. {arch_name_h}</div>
      <div class="muted" style="font-size: 13px; margin-bottom: 16px;">{profile.get("position", "Architetto")}</div>
      <div class="profile-row"><span class="label">Studio</span><span class="value">{arch_company_h}</span></div>
      <div class="profile-row"><span class="label">Ordine</span><span class="value">{arch_ordine_h}</span></div>
      {f'<div class="profile-row"><span class="label">Matricola</span><span class="value">n. {arch_matricola_h}</span></div>' if arch_matricola_h else ''}
      <div class="profile-row"><span class="label">Sede</span><span class="value" style="font-size: 12px;">{arch_address_h}</span></div>
      {f'<div class="profile-row"><span class="label">Email</span><span class="value" style="font-size: 12px;">{arch_email_h}</span></div>' if arch_email_h else ''}
      {f'<div class="profile-row"><span class="label">Tel</span><span class="value">{arch_phone_h}</span></div>' if arch_phone_h else ''}
      {f'<div class="profile-row"><span class="label">PEC</span><span class="value" style="font-size: 11px;">{arch_pec_h}</span></div>' if arch_pec_h else ''}
    </div>
    <div>
      {f'<div class="quote-card" style="margin-bottom: 16px;"><p class="quote" style="font-size: 18px;">{brand.get("mission", "")[:400]}</p><div class="source">— Mission · {brand.get("name", arch_company_h)}</div></div>' if brand.get("mission") else ''}
      {f'<div class="quote-card" style="margin-bottom: 16px; background: linear-gradient(135deg, rgba(135,164,122,0.08) 0%, transparent 100%); border-left-color: #87A47A;"><p class="quote" style="font-size: 16px;">{brand.get("vision", "")[:400]}</p><div class="source">— Vision</div></div>' if brand.get("vision") else ''}
      {f'<div style="background: var(--card); padding: 20px; border-radius: 12px; border: 1px solid var(--border);"><div style="font-family: Outfit; font-weight: 600; margin-bottom: 12px;">Valori dello studio</div><div style="font-size: 13px; line-height: 1.7; color: var(--text-muted); white-space: pre-line;">{brand.get("values", "")}</div></div>' if brand.get("values") else ''}
    </div>
  </div>

  {('<div style="margin-top: 32px;"><div style="font-family: Outfit; font-weight: 600; margin-bottom: 12px;">Paleta della marca</div><div style="display: flex; gap: 12px; flex-wrap: wrap;">' +
    "".join(f'<div style="text-align: center;"><div style="width: 60px; height: 60px; border-radius: 8px; background: {c}; box-shadow: inset 0 0 0 1px rgba(0,0,0,0.05); margin-bottom: 6px;"></div><div style="font-size: 11px; font-family: DM Sans; color: var(--text-muted);">{c}</div></div>'
            for c in (brand.get("palette") or [])) + '</div></div>') if brand.get("palette") else ''}

  {f'<div style="margin-top: 32px;"><div style="font-family: Outfit; font-weight: 600; margin-bottom: 12px;">Stile architettonico dello studio</div><div style="background: var(--card); padding: 16px; border-radius: 10px; border: 1px solid var(--border);"><strong>{style_data.get("style_name", "")}</strong><div style="font-size: 13px; color: var(--text-muted); margin-top: 6px;">Keywords: {", ".join(style_data.get("keywords", []))}</div><div style="font-size: 13px; color: var(--text-muted); margin-top: 4px;">Materiali firma: {", ".join(style_data.get("materials", []))}</div></div></div>' if style_data.get("style_name") else ''}
</section>

<footer>
  <div class="inner">
    <h3>Arch. {arch_name_h}{f" · {arch_company_h}" if arch_company_h else ""}</h3>
    <p style="opacity: 0.85;">{arch_ordine_h}{f" · n. matr. {arch_matricola_h}" if arch_matricola_h else ""}{f" · {arch_address_h}" if arch_address_h else ""}</p>
    <p style="opacity: 0.75; font-size: 13px;">{arch_email_h}{f" · {arch_phone_h}" if arch_phone_h else ""}{f" · PEC {arch_pec_h}" if arch_pec_h else ""}</p>
    <div class="links">
      <span>Contratto firmato 24/04/2026</span>
      <span>Onorari € 22.000 · 4 SAL</span>
      <span>Generato dallo squad architettura-progetto v2.0</span>
      <span>Lovarch Platform · 25/04/2026</span>
    </div>
  </div>
</footer>

</body></html>"""


# ============================================================================
# JSON · text/plain (bucket doesn't allow application/json)
# ============================================================================

def gen_json_pretty(data: Any) -> bytes:
    return json.dumps(data, indent=2, ensure_ascii=False).encode("utf-8")


# ============================================================================
# ZIP · bundle DOSSIER-IMPRESA.zip
# ============================================================================

def gen_zip_dossier(files: List[Tuple[str, bytes]]) -> bytes:
    """Bundle multiple files into a zip · for impresa."""
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for name, content in files:
            z.writestr(name, content)
    return buf.getvalue()


# ============================================================================
# MIME types map · matches user-assets bucket allowlist
# ============================================================================

MIME_BY_EXT = {
    ".pdf": "application/pdf",
    ".dxf": "application/acad",
    ".dwg": "application/dwg",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".zip": "application/zip",
    ".json": "application/json",
    ".html": "text/html",
    ".htm": "text/html",
    ".txt": "text/plain",
    ".csv": "text/csv",
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
}


def mime_for(filename: str) -> str:
    ext = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return MIME_BY_EXT.get(ext, "application/octet-stream")
